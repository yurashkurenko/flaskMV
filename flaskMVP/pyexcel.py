from openpyxl import Workbook
wb = Workbook()
#import json
# grab the active worksheet
#ws = wb.active

# Data can be assigned directly to cells
#ws['A1'] = 42

# Rows can also be appended
#ws.append([1, 2, 3])

# Python types will automatically be converted
import datetime
#ws['A2'] = datetime.datetime.now()

# Save the file
#wb.save("sample.xlsx")

def excel(req):
#    y = json.loads(req.data)
    wb = Workbook()
    ws = wb.active
    ws['A1'] = req.get('elem_a1')
    ws['C13'] = req.get('elem_c13')
    wb.save("sample.xlsx")